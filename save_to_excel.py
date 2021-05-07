from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

from openpyxl.utils import get_column_letter

medium_border = (Border(left=Side(style='medium'),
                            right=Side(style='medium'),
                            top=Side(style='medium'),
                            bottom=Side(style='medium')))

thin_border = (Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin')))

alignment = Alignment(horizontal='center')


def save_short_data_to_excel(artist_name, data_short):

    wb = Workbook()

    ws = wb.active

    ws['A1'] = '№'
    ws['B1'] = 'Name'
    ws['C1'] = 'ID'
    ws['D1'] = 'Release_Date'
    ws['E1'] = 'Total_Tracks'
    ws['F1'] = 'URL'


    for i, album in enumerate(data_short.keys()):
        _ = ws.cell(row=i+2, column=1, value=i+1)
        _ = ws.cell(row=i+2, column=2, value=data_short[album]['name'])
        _ = ws.cell(row=i+2, column=3, value=album)
        _ = ws.cell(row=i+2, column=4, value=data_short[album]['release_date'])
        _ = ws.cell(row=i+2, column=5, value=data_short[album]['total_tracks'])
        _ = ws.cell(row=i+2, column=6, value=data_short[album]['url'])

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 5
        ws.column_dimensions[column_cells[0].column_letter].width = length


    headers = ws[1]
    for cell in headers:
        cell.font = Font(bold=True)
        cell.border = medium_border

    for row in ws.iter_rows(min_row=2, max_col=6):
        for cell in row:
            cell.border = thin_border


    wb.save('{}_short_discography.xlsx'.format(artist_name))

    return wb

def save_full_data_to_excel(artist_name, data_full):

        wb = save_short_data_to_excel(artist_name, data_full)

        for i, album in enumerate(data_full.keys()):

            ws = wb.create_sheet(str(i+1))

            _ = ws.cell(row=2, column=1, value='№')
            _ = ws.cell(row=3, column=1, value='Name')
            _ = ws.cell(row=4, column=1, value='ID')
            _ = ws.cell(row=5, column=1, value='Release_Date')
            _ = ws.cell(row=6, column=1, value='Total_Tracks')
            _ = ws.cell(row=7, column=1, value='URL')

            _ = ws.cell(row=2, column=2, value=i + 1)
            _ = ws.cell(row=3, column=2, value=data_full[album]['name'])
            _ = ws.cell(row=4, column=2, value=album)
            _ = ws.cell(row=5, column=2, value=data_full[album]['release_date'])
            _ = ws.cell(row=6, column=2, value=data_full[album]['total_tracks'])
            _ = ws.cell(row=7, column=2, value=data_full[album]['url'])

            ws['D1'] = 'Disc_Number'
            ws['E1'] = 'Track_Number'
            ws['F1'] = 'Name'
            ws['G1'] = 'ID'
            ws['H1'] = 'Duration_HH:mm:ss.ms'
            ws['I1'] = 'URL'


            for j, track in enumerate(data_full[album]['tracks'].keys()):
                _ = ws.cell(row=j + 2, column=4, value=data_full[album]['tracks'][track]['disc_number'])
                _ = ws.cell(row=j + 2, column=5, value=data_full[album]['tracks'][track]['track_number'])
                _ = ws.cell(row=j + 2, column=6, value=data_full[album]['tracks'][track]['name'])
                _ = ws.cell(row=j + 2, column=7, value=track)
                _ = ws.cell(row=j + 2, column=8, value=data_full[album]['tracks'][track]['duration'])
                _ = ws.cell(row=j + 2, column=9, value=data_full[album]['tracks'][track]['track_url'])

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells) + 5
                ws.column_dimensions[column_cells[0].column_letter].width = length

            for row in ws.iter_cols(min_row=1, min_col=4, max_col=9, max_row=1):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.border = medium_border


            for row in ws.iter_cols(min_row=2, max_col=1, max_row=7):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.border = medium_border


            for row in ws.iter_cols(min_row=2, min_col=2, max_col=2, max_row=7):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left")

            for row in ws.iter_cols(min_row=2, min_col=4, max_col=9):
                for cell in row:
                    cell.border = thin_border


        wb.save('{}_full_discography.xlsx'.format(artist_name))

